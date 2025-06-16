import os
import re
import matplotlib.pyplot as plt
import matplotlib.colors as mcolors
from matplotlib.ticker import MaxNLocator
import pandas as pd
import openpyxl
from datetime import datetime
from scipy.interpolate import UnivariateSpline
import numpy as np

class PlotGenerator:
    def __init__(self, xlabel="Time", ylabel="Number of buildings", title="Accessibility"):
        """
        Initializes the PlotGenerator with the given parameters.
        
        Parameters:
            xlabel (str): Label for the X-axis.
            ylabel (str): Label for the Y-axis.
            title (str): Title of the plot.
        """
        self.xlabel = xlabel
        self.ylabel = ylabel
        self.title = title
        self.lines = []  # List to store line data
        self.colors = list(mcolors.TABLEAU_COLORS.values())  # Using Tableau color palette
        self.output_lines = []
        self.output_lines.append(f"type,time,jaccard_similarity,new_ratio,removed_ratio,value")
        

    @staticmethod
    def count_rows(df, **kwargs):
        """
        Method to count rows in the DataFrame.

        Parameters:
            df (DataFrame): Input DataFrame.

        Returns:
            int: Number of rows.
        """
        return df.shape[0]

    @staticmethod
    def average_travel_time(df, **kwargs):
        """
        Method to calculate the average travel time in the DataFrame.

        Parameters:
            df (DataFrame): Input DataFrame.

        Returns:
            float: Average travel time.
        """
        if 'Duration' in df.columns:
            return df['Duration'].mean()
        return 0
    
    @staticmethod
    def average_wait_time(df):
        """
        Method to calculate the Average Wait Time for the DataFrame.

        Parameters:
        df (DataFrame): Input DataFrame.

        Returns:
        float: Average Wait Time.
        """
        wait_times = ['Wait_time1', 'Wait_time2', 'Wait_time3']
            
        total_wait_time = df[wait_times].sum().sum()
    
        valid_rows = df[wait_times].notna().sum().sum()

        if valid_rows > 0:
            return total_wait_time / valid_rows
        return 0
    
    @staticmethod
    def average_walk_time(df):
        """
        Method to calculate the Average Walk Time for the DataFrame.

        Parameters:
        df (DataFrame): Input DataFrame.

        Returns:
        float: Average walk Time.
        """
        walk_times = ['Walk_time1', 'Walk_time2', 'Walk_time3', 'DestWalk_time']
        total_walk_time = df[walk_times].sum().sum()
        valid_rows = df[walk_times].notna().sum().sum()

        if valid_rows > 0:
            #print (total_walk_time / valid_rows)
            return total_walk_time / valid_rows
            
        return 0

    @staticmethod
    def total_routes(df):
        """
        Method to calculate the Total Routes considering Line_ID1, Line_ID2, Line_ID3.

        Parameters:
            df (DataFrame): Input DataFrame.

        Returns:
            int: Total Routes.
        """
        melted_df = df[['Line_ID1', 'Line_ID2', 'Line_ID3']].melt(value_name='Line_ID')
        return melted_df['Line_ID'].dropna().nunique()    

    
    @staticmethod
    def calculate_convenience_coefficient(df, 
                                          average_travel_time, 
                                          average_wait_time, 
                                          count_transfers,
                                          average_walk_time):
       
        avg_travel_time = average_travel_time(df)
        avg_wait_time = average_wait_time(df)
        avg_walk_time = average_walk_time(df)
        mean_transfers = df.apply(count_transfers, axis=1).mean()
               
        convenience_coefficient = avg_travel_time / ((mean_transfers + 1)*(avg_wait_time + avg_walk_time))

        return convenience_coefficient
    
    @staticmethod
    def count_transfers(row):
        """
        Calculate the number of transfers for a row.

        Parameters:
            row (Series): Input row.

        Returns:
            int: Number of transfers.
        """
        transfers = -1

        if pd.notna(row['Bus_start_time1']) and pd.notna(row['Bus_finish_time1']):
            transfers += 1
        if pd.notna(row['Bus_start_time2']) and pd.notna(row['Bus_finish_time2']):
            transfers += 1
        if pd.notna(row['Bus_start_time3']) and pd.notna(row['Bus_finish_time3']):
            transfers += 1
        return transfers

    def get_sorted_dirs(self, base_dir):
        dir_pattern = re.compile(r"-(\d+)$")
        # Получаем список всех подкаталогов
        all_dirs = [d for d in os.listdir(base_dir) if os.path.isdir(os.path.join(base_dir, d))]
    
        # Сортируем каталоги, извлекая числовой постфикс
        sorted_dirs = sorted(all_dirs, key=lambda d: int(dir_pattern.search(d).group(1)) if dir_pattern.search(d) else float('inf'))
    
        return sorted_dirs

    def add_line(self, base_dir, 
                    min_transfer = 0, 
                    max_transfer = 2, 
                    color = None, 
                    label = None, 
                    calculation_method = None,
                    average_travel_time = None, 
                    average_wait_time = None, 
                    count_transfers = None,
                    average_walk_time = None ):
        """
        Adds a line to the plot using the log file and the corresponding CSV file in the specified directory.

        Parameters:
            base_dir (str): Path to the directory containing the data for this line.
            min_transfer (int): Minimum number of transfers to filter.
            max_transfer (int): Maximum number of transfers to filter.
            color (str): Color of the line (optional, defaults to a color from the palette).
            label (str): Label for the line (optional, defaults to "Line X").
            calculation_method (callable): Method for calculating Y values (default: count_rows).
        """
        # Use a color from the palette if none is provided
        color = color or self.colors[len(self.lines) % len(self.colors)]
        
        # If no label is provided, use a default label
        label = label or f'Line {len(self.lines) + 1}'

        # Default calculation method
        calculation_method = calculation_method or PlotGenerator.count_rows

        # Pattern to extract time from log files
        #The earliest arrival time
        #Arrive before (hh:mm:ss):
                
        time_pattern = re.compile(
            r"Start at \(hh:mm:ss\):\s+(\d{2}:\d{2}:\d{2})|"
            r"The earliest start time:\s+(\d{2}:\d{2}:\d{2})|"
            r"Arrive before \(hh:mm:ss\):\s+(\d{2}:\d{2}:\d{2})|"
            r"The earliest arrival time:\s+(\d{2}:\d{2}:\d{2})"
            )

        # Prepare to collect the lines of data
        x_values = []
        y_values = []

        prev_ids = set() 
        sorted_dirs = self.get_sorted_dirs(base_dir)
        # Searching the directory for log files
        for dir_name in sorted_dirs:
            dir_path = os.path.join(base_dir, dir_name)
            for root, dirs, files in os.walk(dir_path):
                log_files = [f for f in files if f.startswith("log_") and f.endswith(".txt")]
                if log_files:
                    log_file = log_files[0]  # Take the first log file found
                    log_path = os.path.join(root, log_file)

                    # Read the log file
                    with open(log_path, 'r') as log:
                        log_content = log.read()
                        time_matches = list(time_pattern.finditer(log_content))  # Find all matches
                        if time_matches:
                            last_match = time_matches[-1]  # Take the last match
                            if last_match.group(1):
                                time_str = last_match.group(1)
                            elif last_match.group(2):
                                time_str = last_match.group(2)
                            elif last_match.group(3):
                                time_str = last_match.group(3)    
                            elif last_match.group(4):
                                time_str = last_match.group(4)    
        
                            time_without_seconds = time_str[:5]  # Keep only hh:mm part
                            x_values.append(time_without_seconds)  # Add time to X

                            # Find the corresponding CSV file (assuming only one in the directory)
                            base_name = log_file.replace("log_", "").replace(".txt", "")
                            csv_file_path = None
                            for file in files:
                                if file.startswith(base_name) and file.endswith(".csv"):
                                    csv_file_path = os.path.join(root, file)
                                    break

                            if csv_file_path:
                                # Read the CSV file using pandas
                                df = pd.read_csv(csv_file_path)
                                                      
                                # Apply the transfer conditions
                                if min_transfer == 0 and max_transfer == 2:
                                    filtered_df = df[df['Start_time'].notna()]
                                else:
                                    if 'Transfers' not in df.columns:
                                        df['Transfers'] = df.apply(self.count_transfers, axis=1)
                                        df.to_csv(csv_file_path, index=False)
                                    filtered_df = df[(df['Transfers'] >= min_transfer) & (df['Transfers'] <= max_transfer)]
                                ###########
                                current_ids = set(df['Destination_ID'])

                                # Вычисляем метрику изменения между текущим и предыдущим набором
                                jaccard_similarity = 0
                                new_ratio = 0
                                removed_ratio = 0
                                if prev_ids:
                                    intersection = len(current_ids & prev_ids)
                                    union = len(current_ids | prev_ids)

                                    # Jaccard Similarity
                                    jaccard_similarity = intersection / union if union > 0 else 1

                                    # Доля новых значений
                                    new_ratio = len(current_ids - prev_ids) / len(current_ids) if current_ids else 0

                                    # Доля исчезнувших значений
                                    removed_ratio = len(prev_ids - current_ids) / len(prev_ids) if prev_ids else 0

                                    if removed_ratio == 0:
                                        print (f' len(prev_ids - current_ids) {len(prev_ids - current_ids)} len(prev_ids) {len(prev_ids)}')
                                
                                prev_ids = current_ids    
                            
                                ###########
                                # Calculate the Y value using the provided method
                                # Если calculation_method — это функция для расчёта коэффициента удобства
                                if calculation_method == PlotGenerator.calculate_convenience_coefficient:
                                    y_value = calculation_method(filtered_df, 
                                                             average_travel_time, 
                                                             average_wait_time, 
                                                             count_transfers, 
                                                             average_walk_time
                                                             )
                                else:
                                    y_value = calculation_method(filtered_df)
                                y_values.append(y_value)
                                self.output_lines.append(f"{label},{time_without_seconds},{jaccard_similarity:.2f},{new_ratio:.2f},{removed_ratio:.2f},{y_value}")
                            else:
                                y_values.append(0)
                                print(f"CSV file corresponding to {log_file} not found. {csv_file_path}")

        print(f"len(x_values): {len(x_values)}, len(y_values): {len(y_values)}")
        

        if x_values and y_values:
            datetime_x_values = [datetime.strptime(time, "%H:%M") for time in x_values]
            sorted_indices = sorted(range(len(datetime_x_values)), key=lambda i: datetime_x_values[i])
        
            x_values = [x_values[i] for i in sorted_indices]
            y_values = [y_values[i] for i in sorted_indices]

            self.lines.append((x_values, y_values, color, label))
        else:
            print(f"No data found for directory: {base_dir}.")

    def generate_plot(self, show_legend=True):
        """
        Generates the plot with the added lines.

        Parameters:
            show_legend (bool): Whether to display the legend on the plot.
        """
        plt.figure(figsize=(10, 6))

        # Plot each line
        for idx, (x_values, y_values, color, label) in enumerate(self.lines):
            plt.plot(x_values, y_values, marker='o', linestyle='-', color=color)  # Plot the line

            # Add the label to the plot at the start of the line if legend is not shown
            if not show_legend:
                x_start = x_values[0]  # First X value (start of the line)
                y_start = y_values[0]  # First Y value (start of the line)

                plt.text(x_start, y_start - 0.05, label, color=color, fontsize=10,
                         verticalalignment='bottom', horizontalalignment='left')


            # Add the label to the plot at the start of the line if legend is not shown
            if not show_legend:
                x_start = x_values[0]  # First X value (start of the line)
                y_start = y_values[0]  # First Y value (start of the line)
                plt.text(x_start, y_start - 0.05, label, color=color, fontsize=10,
                     verticalalignment='bottom', horizontalalignment='left')

        if not self.lines:
            print("No lines to plot. Ensure that data is added correctly.")
            return

        if show_legend:
            plt.legend([label for _, _, _, label in self.lines], loc='best', bbox_to_anchor=(1.05, 1), borderaxespad=0.)

        plt.xlabel(self.xlabel)  # X-axis label
        plt.ylabel(self.ylabel)  # Y-axis label
        plt.title(self.title)  # Plot title
        plt.xticks(rotation=0)  # Set X-axis labels to be horizontal
        plt.grid(True)  # Enable grid on the plot

        
        # Automatically reduce the number of ticks on the X-axis
        ax = plt.gca()  # Get the current axis
        ax.xaxis.set_major_locator(MaxNLocator(integer=True, prune='both', nbins=10))  # Limit to 10 ticks or less
        plt.tight_layout()  # Adjust layout to prevent overlap
        plt.show()
        

    def save_data_to_excel(self, filename):
        """
        Saves the plot data to an Excel file in the specified format.

        Parameters:
            filename (str): Path to the Excel file to save.
        """
        # Initialize the data structure with the 'Time' column
        data = {'Time': []}  # Header for time
        labels = []  # List to store unique labels (dynamically filled)
    
        # Fill the data from self.lines
        for idx, (x_values, y_values, _, label) in enumerate(self.lines):
            if label not in labels:  # Add unique labels
                labels.append(label)
                data[label] = []  # Add a column for each label

            # Iterate through the x_values and y_values and add them to the structure
            for i, x in enumerate(x_values):
                if x not in data['Time']:  # Add unique time values
                    data['Time'].append(x)
        
                # For each time, add corresponding values in the respective column
                while len(data[label]) < len(data['Time']):
                    data[label].append(None)  # Fill missing values with None
        
                data[label][data['Time'].index(x)] = y_values[i]

        # Ensure all columns have the same length
        max_length = len(data['Time'])
        for label in labels:
            while len(data[label]) < max_length:
                data[label].append(None)

        # Convert the data structure to a DataFrame
        df = pd.DataFrame(data)

        # Save the DataFrame to an Excel file (without the title)
        df.to_excel(filename, index=False)

        # Open the Excel file to insert formulas for ave, std, and CV
        
        wb = openpyxl.load_workbook(filename)
        ws = wb.active

        # Insert the title in the first row (self.title)
        ws.insert_rows(1)
        ws['A1'] = self.title

        # Get the row number for the first data row (after title)
        first_data_row = 3

        # Add formulas for ave, std, and CV for each label column (starting from column B)
        for col_idx, label in enumerate(labels, start=2):  # Starting from column 2 (for 'Time' column)
            # Get the last row of data (excluding the title row)
            last_row = first_data_row + len(df) - 1

            # The first row for statistics (below data)
            stats_row = last_row + 1
                    
            # Add labels for statistics in column A
            ws[f"A{stats_row}"] = "ave"
            ws[f"A{stats_row+1}"] = "std"
            ws[f"A{stats_row+2}"] = "cv"

            # Formula for average (ave) for each column dynamically, rounded to 1 decimal place
            ave_formula = f"=ROUND(AVERAGE({openpyxl.utils.get_column_letter(col_idx)}{first_data_row}:{openpyxl.utils.get_column_letter(col_idx)}{last_row}), 1)"
            ws[f"{openpyxl.utils.get_column_letter(col_idx)}{stats_row}"] = ave_formula
        
            # Formula for standard deviation (std) using Russian formula for local versions of Excel, rounded to 1 decimal place
            std_formula = f"=ROUND(STDEV({openpyxl.utils.get_column_letter(col_idx)}{first_data_row}:{openpyxl.utils.get_column_letter(col_idx)}{last_row}), 3)"  # Formula for Russian version
            ws[f"{openpyxl.utils.get_column_letter(col_idx)}{stats_row+1}"] = std_formula
        
            # Formula for coefficient of variation (CV), rounded to 1 decimal place
            cv_formula = f"=IFERROR(ROUND({openpyxl.utils.get_column_letter(col_idx)}{stats_row+1}/{openpyxl.utils.get_column_letter(col_idx)}{stats_row}, 3), 1)"
            ws[f"{openpyxl.utils.get_column_letter(col_idx)}{stats_row+2}"] = cv_formula


        # Save the workbook with formulas
        wb.save(filename)

        
        filename_csv = os.path.splitext(filename)[0] + ".csv"
      
        with open(filename_csv, "w") as file:
            file.write("\n".join(self.output_lines))

        print(f"Data successfully saved to {filename}")



if __name__ == "__main__":
    plot_generator = PlotGenerator(
        xlabel="Start time",  # X-axis label
        ylabel="Number of buildings",  # Y-axis label
        title="Accessibility"  # Plot title
    )

    # Add lines from the specified directory with specific transfer counts
    plot_generator.add_line(
        base_dir=r"C:\\Users\\geosimlab\\Documents\\Igor\\experiments\\gesher_from_fix-time\\",        
        min_transfer = 0, max_transfer = 0,
        label="Gesher (from, fix-time)"  # Set line label
    )

    plot_generator.add_line(
        base_dir=r"C:\\Users\\geosimlab\\Documents\\Igor\\experiments\\gesher_from_fix-time\\",        
        min_transfer = 1, max_transfer = 1,
        label="Gesher (from, fix-time)"  # Set line label
    )

    plot_generator.add_line(
        base_dir=r"C:\\Users\\geosimlab\\Documents\\Igor\\experiments\\gesher_from_fix-time\\",        
        min_transfer = 2, max_transfer = 2,
        label="Gesher (from, fix-time)"  # Set line label
    )

    plot_generator.add_line(
        base_dir=r"C:\\Users\\geosimlab\\Documents\\Igor\\experiments\\gesher_from_fix-time\\",        
        min_transfer = 0, max_transfer = 2,
        label="Gesher (from, fix-time)"  # Set line label
    )
    ##################################

    plot_generator.add_line(
        base_dir=r"C:\\Users\\geosimlab\\Documents\\Igor\\experiments\\gesher_from_shedule-base\\",        
        min_transfer = 0, max_transfer = 0,
        label="Gesher (from, shedule-based)"  # Set line label
    )

    plot_generator.add_line(
        base_dir=r"C:\\Users\\geosimlab\\Documents\\Igor\\experiments\\gesher_from_shedule-base\\",        
        min_transfer = 1, max_transfer = 1,
        label="Gesher (from, shedule-based)"  # Set line label
    )

    plot_generator.add_line(
        base_dir=r"C:\\Users\\geosimlab\\Documents\\Igor\\experiments\\gesher_from_shedule-base\\",        
        min_transfer = 2, max_transfer = 2,
        label="Gesher (from, shedule-based)"  # Set line label
    )

    plot_generator.add_line(
        base_dir=r"C:\\Users\\geosimlab\\Documents\\Igor\\experiments\\gesher_from_shedule-base\\",        
        min_transfer = 0, max_transfer = 2,
        label="Gesher (from, shedule-based)"  # Set line label
    )

    # Save data to an Excel file
    plot_generator.save_data_to_excel(r"C:\\Users\\geosimlab\\Documents\\Igor\\experiments\\plot_data.xlsx")

    # Generate the plot
    plot_generator.generate_plot()
